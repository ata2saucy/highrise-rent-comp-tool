#!/usr/bin/env python3
"""
build_workbook.py — deterministic v2 rent-comp workbook generator.

The agent's job is to assemble a verified `comp_data.json` (see comp_data.schema.json
and comp_data.example.json). THIS script owns 100% of the workbook's structure and
formatting, so the output is identical every run and never "chopped". Do NOT hand-build
the workbook — run this.

    python build_workbook.py comp_data.json ["Out Name Rental Comps _vACTIVE.xlsx"]

Produces the canonical 6 sheets, in order:
  Output · Building Summary · Data_Summary · RD Condos · RD Apartments · Floor Plans

Formatting (widths, number formats, fills, fonts, frozen panes, hidden gridlines,
fullCalcOnLoad) is defined here in code — it is the single source of truth for the v2
format. Levers per the current CLAUDE.md: C2 = hard-coded parking adjustment + source
note (NO LINEST / no J:N helper block); C4 = live formula off PREMIUM BASIS.
"""
import sys, json, datetime, os, shutil, subprocess, tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styles
NAVY   = "FF1F3864"   # header band
WHITE  = "FFFFFFFF"
BLUE   = "FF0000FF"   # hard-coded input font
GREY   = "FF595959"   # note font
GREEN_F= "FF1B5E20"   # recommendation font
LBLUE  = "FFCCECFF"   # group/label fill
CREAM_L= "FFFFF2CC"   # subject-site / cream lever
BAND   = "FFD9E1F2"   # ALL-band fill
RECG   = "FFC6EFCE"   # recommended cell fill
YEL    = "FFFFFF00"   # tunable lever fill
GREEN  = "FFE2EFDA"   # confidence: leased+verified
CREAM  = "FFFFF2CC"   # confidence: active / older / caveat / apartment asking
ORANGE = "FFFCE4D6"   # confidence: excluded partial
REG    = "FFFFFFCC"   # regression note fill

F_HDR  = Font(bold=True, color=WHITE)
F_TITLE= Font(bold=True, color=NAVY)
F_BOLD = Font(bold=True)
F_BLUE = Font(color=BLUE)
F_GREY = Font(color=GREY)
F_REC  = Font(bold=True, color=GREEN_F)
FILL_HDR  = PatternFill("solid", fgColor=NAVY)
FILL_LBLUE= PatternFill("solid", fgColor=LBLUE)
FILL_CREAM= PatternFill("solid", fgColor=CREAM_L)
FILL_BAND = PatternFill("solid", fgColor=BAND)
FILL_REC  = PatternFill("solid", fgColor=RECG)
FILL_YEL  = PatternFill("solid", fgColor=YEL)
FILL_REG  = PatternFill("solid", fgColor=REG)
CONF_FILL = {"green": PatternFill("solid", fgColor=GREEN),
             "cream": PatternFill("solid", fgColor=CREAM),
             "orange":PatternFill("solid", fgColor=ORANGE)}

# number formats
NF_INT  = "#,##0"
NF_PSF  = "#,##0.000"
NF_DATE = "yyyy\\-mm\\-dd"
NF_PCT  = "0.0%"
NF_USD  = "\\$#,##0"
NF_USD2 = "\\$#,##0.00"
NF_PSF2 = "0.000"

def C(ws, addr, value=None, nf=None, font=None, fill=None, align=None):
    c = ws[addr]
    if value is not None: c.value = value
    if nf: c.number_format = nf
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    return c

def hdr(ws, addr, text):
    C(ws, addr, text, font=F_HDR, fill=FILL_HDR)

def widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w

def autofit(ws, min_w=6, max_w=100, pad=1):
    """Size each column to its longest cell's displayed text (one line, no wrap). Formula cells
    are skipped — Excel computes them at open, so their width can't be measured here; the
    column's header / literal cells drive the width instead. Multi-line headers (with '\\n') are
    measured per line. Capped at max_w so a long free-text field (description / URL / amenities)
    can't blow a column out to hundreds of characters."""
    best = {}
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if isinstance(v, str):
                if v.startswith("="):                       # formula → measured at open, skip
                    continue
                ln = max((len(s) for s in v.split("\n")), default=0)
            elif isinstance(v, bool):
                ln = 5
            elif isinstance(v, (datetime.datetime, datetime.date)):
                ln = 10                                      # yyyy-mm-dd
            elif isinstance(v, float):
                ln = len(f"{v:,.3f}")
            elif isinstance(v, int):
                ln = len(f"{v:,}")
            else:
                ln = len(str(v))
            col = c.column_letter
            if ln > best.get(col, 0):
                best[col] = ln
    for col, ln in best.items():
        ws.column_dimensions[col].width = max(min_w, min(max_w, ln + pad))

def dt(s):
    if isinstance(s, str) and len(s) == 10 and s[4] == "-":
        try: return datetime.datetime.strptime(s, "%Y-%m-%d")
        except ValueError: return s
    return s

# ---------------------------------------------------------------- RD sheets
RD_HEADERS = ["Include","Sq Ft.","Rent","Adj. Rent","$/sq ft","Adj. $/sq ft","Adj BD",
 "Beds Number","Den","Date","Building Name","Building Address","Building City","Developer",
 "Unit #","Unit Address","Beds","Baths","Sqft (Condos.ca)","MLS Size Range","# Parking",
 "Parking Included","Locker","Outdoor Space","Exposure","Building Age","Building Amenities",
 "Leased Rent ($)","Listed Rent ($)","PSF (Calculated)","PSF (Listing)","Furnished",
 "Hydro Included","Water Included","Lease Date","MLS#","Listing URL","Description",
 "Product Type","$/SF (Product-Adj)","Date Scraped"]   # 41 cols A:AO
RD_W = {"A":8,"B":8,"C":9,"D":9,"E":9,"F":9,"G":7,"H":7,"I":6,"J":12,"K":24,"L":22,"M":18,
 "N":16,"O":9,"P":26,"Q":8,"R":7,"S":11,"T":16,"U":8,"V":9,"W":12,"X":18,"Y":9,"Z":18,
 "AA":40,"AB":12,"AC":12,"AD":11,"AE":11,"AF":9,"AG":9,"AH":9,"AI":12,"AJ":12,"AK":60,
 "AL":60,"AM":12,"AN":12,"AO":12}

def build_rd(wb, title, rows, product):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    widths(ws, RD_W)
    for j, h in enumerate(RD_HEADERS, start=1):
        C(ws, f"{get_column_letter(j)}1", h, font=F_HDR, fill=FILL_HDR)
    ws.freeze_panes = "A2"
    for i, rec in enumerate(rows):
        r = i + 2
        # live formulas
        C(ws, f"A{r}", f"=IF(AND(ISNUMBER(B{r}),B{r}>0,ISNUMBER(AB{r}),AB{r}>0,J{r}>=Data_Summary!$C$1),1,0)")
        C(ws, f"C{r}", f"=IF(ISNUMBER(AB{r}),AB{r},\"\")", nf=NF_INT)
        C(ws, f"D{r}", f"=IF(ISNUMBER(C{r}),C{r}-U{r}*Data_Summary!$C$2,\"\")")
        C(ws, f"E{r}", f"=IFERROR(C{r}/B{r},\"\")", nf=NF_PSF)
        C(ws, f"F{r}", f"=IFERROR(D{r}/B{r},\"\")", nf=NF_PSF)
        C(ws, f"G{r}", f"=Q{r}")
        C(ws, f"H{r}", f"=IFERROR(LEFT(G{r},1),\"\")")
        C(ws, f"I{r}", f"=IF(LEN(G{r})>1,1,0)")
        C(ws, f"AD{r}", f"=IFERROR(AB{r}/B{r},\"\")", nf=NF_PSF)
        C(ws, f"AE{r}", f"=IFERROR(AC{r}/B{r},\"\")", nf=NF_PSF)
        C(ws, f"AN{r}", f"=IFERROR(E{r}*(1+IF(AM{r}=\"Apartment\",Data_Summary!$C$3,0)),\"\")", nf=NF_PSF)
        # data (hard) columns
        C(ws, f"B{r}", rec.get("sqft"), nf=NF_INT)
        C(ws, f"J{r}", dt(rec.get("date")), nf=NF_DATE)
        C(ws, f"K{r}", rec.get("building"))
        C(ws, f"L{r}", rec.get("building_address"))
        C(ws, f"M{r}", rec.get("building_city"))
        C(ws, f"N{r}", rec.get("developer"))
        C(ws, f"O{r}", rec.get("unit"))
        C(ws, f"P{r}", rec.get("unit_address"))
        C(ws, f"Q{r}", rec.get("beds"))
        C(ws, f"R{r}", rec.get("baths"))
        C(ws, f"S{r}", rec.get("sqft_condosca"), nf=NF_INT)
        C(ws, f"T{r}", rec.get("mls_range"))
        C(ws, f"U{r}", rec.get("parking"))
        C(ws, f"V{r}", rec.get("parking_incl"))
        C(ws, f"W{r}", rec.get("locker"))
        C(ws, f"X{r}", rec.get("outdoor"))
        C(ws, f"Y{r}", rec.get("exposure"))
        C(ws, f"Z{r}", rec.get("building_age"))
        C(ws, f"AA{r}", rec.get("amenities"))
        C(ws, f"AB{r}", rec.get("leased_rent"), nf=NF_INT, font=F_BLUE)
        C(ws, f"AC{r}", rec.get("listed_rent"), nf=NF_INT, font=F_BLUE)
        C(ws, f"AF{r}", rec.get("furnished"))
        C(ws, f"AG{r}", rec.get("hydro"))
        C(ws, f"AH{r}", rec.get("water"))
        C(ws, f"AI{r}", rec.get("lease_date"))
        C(ws, f"AJ{r}", rec.get("mls"))
        C(ws, f"AK{r}", rec.get("url"))
        # Description: prefer the supplied desc; else, for a documented SF blank, surface the
        # exhaustion record (condos.ca page + key-plate + outcome) so the provenance is visible.
        desc = rec.get("desc")
        sb = rec.get("sf_blank")
        if not desc and isinstance(sb, dict):
            desc = (f"SF BLANK [{sb.get('outcome','')}] — "
                    f"condos.ca page: {sb.get('condosca_page','')}; "
                    f"key-plate: {sb.get('keyplate','')}")
        C(ws, f"AL{r}", desc)
        C(ws, f"AM{r}", rec.get("product_type", product))
        C(ws, f"AO{r}", dt(rec.get("date_scraped")), nf=NF_DATE)
        # confidence fills on B & E
        conf = rec.get("confidence")
        if conf in CONF_FILL:
            ws[f"B{r}"].fill = CONF_FILL[conf]
            ws[f"E{r}"].fill = CONF_FILL[conf]
    if not rows:
        C(ws, "K3", f"No purpose-built rental apartment comps in this set."
                    if product == "Apartment" else "No rows.", font=F_GREY)
    return ws

# ---------------------------------------------------------------- Floor Plans
FP_HEADERS = ["Building Name","Building Address","Building City","Subject/Comp","Source Site",
 "Source URL","Date Read","Suite / Plan Name","Beds","Baths","Interior SF","Exposure",
 "Floor Band","Stack / Line","Balcony/Terrace","Notes","Used For Unit(s)","Verification Tier"]
FP_W = {"A":34,"B":22,"C":22,"D":12,"E":18,"F":55,"G":12,"H":18,"I":6,"J":6,"K":11,"L":10,
 "M":22,"N":10,"O":14,"P":55,"Q":16,"R":16}
FP_KEYS = ["building","address","city","role","source_site","source_url","date_read",
 "plan_name","beds","baths","interior_sf","exposure","floor_band","stack","balcony",
 "notes","used_for","tier"]

def build_floor_plans(wb, plans, note):
    ws = wb.create_sheet("Floor Plans")
    ws.sheet_view.showGridLines = False
    widths(ws, FP_W)
    for j, h in enumerate(FP_HEADERS, start=1):
        C(ws, f"{get_column_letter(j)}1", h, font=F_BOLD)
    ws.freeze_panes = "A2"
    r = 2
    for rec in plans:
        for j, k in enumerate(FP_KEYS, start=1):
            v = rec.get(k)
            cell = C(ws, f"{get_column_letter(j)}{r}", v)
            if k == "interior_sf" and v is not None:
                cell.number_format = NF_INT
        r += 1
    if note:
        C(ws, f"A{r}", note, font=F_GREY)
    return ws

# ---------------------------------------------------------------- Building Summary
BS_HEADERS = ["Building","Address","Area","Developer","First Occupancy","Yr Built","Product",
 "Storeys","Units","Transaction Count","Avg SF","Avg Rent","Avg $/SF"]
BS_W = {"A":24,"B":24,"C":18,"D":16,"E":14,"F":9,"G":9,"H":8,"I":8,"J":14,"K":9,"L":10,"M":10}

def build_building_summary(wb, subject, buildings):
    ws = wb.create_sheet("Building Summary")
    ws.sheet_view.showGridLines = False
    widths(ws, BS_W)
    C(ws, "A1", f"Comp Building Summary — {subject['name']} "
                f"({subject['address']}, {subject.get('area_label','')})", font=F_TITLE)
    for j, h in enumerate(BS_HEADERS, start=1):
        C(ws, f"{get_column_letter(j)}3", h, font=F_HDR, fill=FILL_HDR)
    r0 = 4
    for i, b in enumerate(buildings):
        r = r0 + i
        C(ws, f"A{r}", b.get("name"))
        C(ws, f"B{r}", b.get("address"))
        C(ws, f"C{r}", b.get("area"))
        C(ws, f"D{r}", b.get("developer"))
        C(ws, f"E{r}", b.get("first_occ"))
        C(ws, f"F{r}", b.get("yr_built"))
        C(ws, f"G{r}", b.get("product"))
        C(ws, f"H{r}", b.get("storeys"))
        C(ws, f"I{r}", b.get("units"))
        C(ws, f"J{r}", f"=COUNTIFS('RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1)")
        C(ws, f"K{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$B:$B,'RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"L{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$AB:$AB,'RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"M{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$E:$E,'RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1),\"\")",
          nf=NF_PSF, fill=FILL_REC)
    nb = len(buildings)
    rc = r0 + nb         # ALL CONDOS
    ra = r0 + nb + 1     # ALL APARTMENTS
    C(ws, f"A{rc}", "ALL CONDOS", fill=FILL_BAND)
    C(ws, f"J{rc}", "=SUM(J%d:J%d)" % (r0, r0+nb-1), fill=FILL_BAND)
    C(ws, f"K{rc}", "=IFERROR(AVERAGEIFS('RD Condos'!$B:$B,'RD Condos'!$A:$A,1),\"\")", nf=NF_INT, fill=FILL_BAND)
    C(ws, f"L{rc}", "=IFERROR(AVERAGEIFS('RD Condos'!$AB:$AB,'RD Condos'!$A:$A,1),\"\")", nf=NF_INT, fill=FILL_BAND)
    C(ws, f"M{rc}", "=IFERROR(AVERAGEIFS('RD Condos'!$AN:$AN,'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF, fill=FILL_BAND)
    C(ws, f"A{ra}", "ALL APARTMENTS", fill=FILL_BAND)
    C(ws, f"J{ra}", "=COUNTIFS('RD Apartments'!$A:$A,1)", fill=FILL_BAND)
    C(ws, f"K{ra}", "=IFERROR(AVERAGEIFS('RD Apartments'!$B:$B,'RD Apartments'!$A:$A,1),\"\")", nf=NF_INT, fill=FILL_BAND)
    C(ws, f"L{ra}", "=IFERROR(AVERAGEIFS('RD Apartments'!$AB:$AB,'RD Apartments'!$A:$A,1),\"\")", nf=NF_INT, fill=FILL_BAND)
    C(ws, f"M{ra}", "=IFERROR(AVERAGEIFS('RD Apartments'!$AN:$AN,'RD Apartments'!$A:$A,1),\"\")", nf=NF_PSF, fill=FILL_BAND)
    return ws

# ---------------------------------------------------------------- Data_Summary
def build_data_summary(wb, d):
    ws = wb.create_sheet("Data_Summary")
    ws.sheet_view.showGridLines = False
    widths(ws, {"A":42,"B":22,"C":14,"D":70,"E":12,"F":14,"G":10,"H":10,
                "P":26,"Q":10,"S":34,"T":12,"U":40})
    subj = d["subject"]; lev = d["levers"]
    primary = subj["primary_building"]; oldest = subj.get("oldest_building", primary)
    pa = lev["parking_adj"]; ap = lev["apt_premium"]
    smo = lev["subject_mix_output"]; smm = lev["subject_mix_manual"]

    # ---- levers C1:C4
    C(ws, "B1", "Date Filter (include leases on/after)")
    C(ws, "C1", dt(lev["date_filter"]), nf=NF_DATE, font=F_BLUE, fill=FILL_YEL)
    C(ws, "D1", "lever — trailing window; leases on/after this date count. Tunable.", font=F_GREY)
    C(ws, "B2", "Parking Adjustment ($/spot/mo)")
    C(ws, "C2", pa["value"], nf=NF_USD2, font=F_BLUE, fill=FILL_YEL)
    C(ws, "D2", pa["note"], font=F_GREY)
    C(ws, "B3", "Apartment premium (apt comps only)" + (" — OFF" if ap.get("off") else ""))
    C(ws, "C3", ap["value"], nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)
    C(ws, "D3", ap["note"], font=F_GREY)
    sp = lev.get("subject_premium")
    fb = sp["value"] if sp else 0
    C(ws, "B4", "Subject rental premium (new-build / lease-up; comp $/SF → subject)")
    # C4 stays a live formula off PREMIUM BASIS = midpoint of the two observed vintage premiums
    # (C61 newest/all-comp, C62 newest/oldest). A single-/equal-vintage comp set has NO spread
    # (C61=C62=0), so it falls back to the sourced premium in levers.subject_premium (else 0) —
    # never a silent 0 that guts the recommendation.
    C(ws, "C4", f"=IF(AND(ABS(C61)<0.0001,ABS(C62)<0.0001),{fb},(C61+C62)/2)", nf=NF_PCT)
    if sp:
        C(ws, "D4", f"DERIVED off PREMIUM BASIS (midpoint of C61 newest/all-comp & C62 "
                    f"newest/oldest). This comp set is single-/equal-vintage ⇒ no spread ⇒ C4 "
                    f"falls back to the sourced subject premium {fb:.1%}: {sp.get('source','')}", font=F_GREY)
    else:
        C(ws, "D4", "DERIVED — midpoint of the two observed vintage premiums in PREMIUM BASIS "
                    "(C61 newest/all-comp, C62 newest/oldest). A single-/equal-vintage comp set ⇒ "
                    "0; supply levers.subject_premium to carry a sourced premium. Not hard-coded.", font=F_GREY)

    # ---- subject unit-mix derivation (H2:H4 + P/Q) and S/T output mix
    C(ws, "G1", "Subject unit mix (1BR/2BR/3BR, den-incl) — DERIVED")
    C(ws, "G2", "1BR wt"); C(ws, "H2", "=IF($Q$5>0,$Q$2/$Q$5,$Q$8)", nf=NF_PCT)
    C(ws, "G3", "2BR wt"); C(ws, "H3", "=IF($Q$5>0,$Q$3/$Q$5,$Q$9)", nf=NF_PCT)
    C(ws, "G4", "3BR wt"); C(ws, "H4", "=IF($Q$5>0,$Q$4/$Q$5,$Q$10)", nf=NF_PCT)
    C(ws, "P1", "Subject Unit-Mix Derivation (P/Q)", font=F_TITLE)
    C(ws, "P2", "1BR plans (SUBJECT, den-incl)"); C(ws, "Q2", "=COUNTIFS('Floor Plans'!$A:$A,\"*(SUBJECT)*\",'Floor Plans'!$I:$I,\"<=1\")")
    C(ws, "P3", "2BR plans (SUBJECT)");            C(ws, "Q3", "=COUNTIFS('Floor Plans'!$A:$A,\"*(SUBJECT)*\",'Floor Plans'!$I:$I,2)")
    C(ws, "P4", "3BR plans (SUBJECT)");            C(ws, "Q4", "=COUNTIFS('Floor Plans'!$A:$A,\"*(SUBJECT)*\",'Floor Plans'!$I:$I,\">=3\")")
    C(ws, "P5", "Total tagged plans");             C(ws, "Q5", "=SUM(Q2:Q4)")
    C(ws, "P6", "(0 ⇒ falls back to manual mix Q8:Q10)", font=F_GREY)
    C(ws, "P7", "Source"); C(ws, "Q7", smm.get("source",""), font=F_BLUE)
    C(ws, "P8", "1BR wt (manual)"); C(ws, "Q8", smm.get("1BR"), nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)
    C(ws, "P9", "2BR wt (manual)"); C(ws, "Q9", smm.get("2BR"), nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)
    C(ws, "P10","3BR wt (manual)"); C(ws, "Q10",smm.get("3BR"), nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)
    C(ws, "S1", "Subject Output mix & suite count", font=F_TITLE)
    C(ws, "S2", "Suite count"); C(ws, "T2", smo.get("suite_count"), font=F_BLUE)
    C(ws, "U2", smo.get("suite_count_note",""), font=F_GREY)
    C(ws, "S3", "1BR wt");   C(ws, "T3", smo.get("1BR"),   nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)
    C(ws, "S4", "1+Den wt"); C(ws, "T4", smo.get("1+Den"), nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)
    C(ws, "S5", "2BR wt");   C(ws, "T5", smo.get("2BR"),   nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)
    C(ws, "S6", "3BR wt");   C(ws, "T6", smo.get("3BR"),   nf=NF_PCT, font=F_BLUE, fill=FILL_YEL)

    # ---- By Bedroom — all comps (rows 18-22)
    C(ws, "A18", "By Bedroom — all comps (den-incl.)", font=F_BOLD)
    for col, t in zip("ABCDE", ["Beds","Incl.","Avg SF","Avg Rent","Avg $/SF"]):
        C(ws, f"{col}19", t, font=F_HDR, fill=FILL_HDR)
    for r, bed in [(20,"1"),(21,"2"),(22,"3")]:
        C(ws, f"A{r}", f"{bed} Bed")
        C(ws, f"B{r}", f"=COUNTIFS('RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1)")
        C(ws, f"C{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$B:$B,'RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"D{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$AB:$AB,'RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"E{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$AN:$AN,'RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF)

    # ---- By Bedroom — primary only (rows 24-28)
    C(ws, "A24", f"By Bedroom — {primary} only (new-build basis)", font=F_BOLD)
    for col, t in zip("ABCDEF", ["Beds","Incl.","Avg SF","Avg Rent","$/SF","Subj Rent (+prem)"]):
        C(ws, f"{col}25", t, font=F_HDR, fill=FILL_HDR)
    for r, bed in [(26,"1"),(27,"2"),(28,"3")]:
        C(ws, f"A{r}", f"{bed} Bed")
        C(ws, f"B{r}", f"=COUNTIFS('RD Condos'!$K:$K,\"{primary}\",'RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1)")
        C(ws, f"C{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$B:$B,'RD Condos'!$K:$K,\"{primary}\",'RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"D{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$AB:$AB,'RD Condos'!$K:$K,\"{primary}\",'RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"E{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$E:$E,'RD Condos'!$K:$K,\"{primary}\",'RD Condos'!$H:$H,\"{bed}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF)
        C(ws, f"F{r}", f"=IFERROR(D{r}*(1+$C$4),\"\")", nf=NF_USD, font=F_REC)

    # ---- SUBJECT RECOMMENDATION (rows 30-37)
    C(ws, "A30", "SUBJECT RECOMMENDATION (comp $/SF + premium)", font=F_REC)
    C(ws, "A31", f"Weighted comp $/SF — {primary} basis"); C(ws, "C31", "=E26*$H$2+E27*$H$3+E28*$H$4", nf=NF_USD2)
    C(ws, "A32", "Weighted comp $/SF — all-comp basis");   C(ws, "C32", "=E20*$H$2+E21*$H$3+E22*$H$4", nf=NF_USD2)
    C(ws, "A33", "Subject rental premium");                C(ws, "C33", "=C4", nf=NF_PCT)
    C(ws, "A34", f"RECOMMENDED subject $/SF ({primary} + premium)")
    C(ws, "C34", "=C31*(1+C4)", nf=NF_USD2, font=F_REC, fill=FILL_REC)
    C(ws, "A35", "Subject $/SF (all-comp + premium)");     C(ws, "C35", "=C32*(1+C4)", nf=NF_USD2)
    C(ws, "A36", "Prior model subject")
    C(ws, "C36", subj.get("prior_model") or "n/a — new subject (no prior model)",
      font=(F_BLUE if subj.get("prior_model") else F_GREY))
    C(ws, "A37", "Residual vs prior")
    C(ws, "C37", "=IFERROR(C34/C36-1,\"\")" if subj.get("prior_model") else "n/a",
      nf=(NF_PCT if subj.get("prior_model") else None), font=(None if subj.get("prior_model") else F_GREY))

    # ---- PREMIUM BASIS (rows 56-63)
    C(ws, "A56", "PREMIUM BASIS (live) — vintage $/SF spread anchoring C4", font=F_BOLD)
    C(ws, "A57", "Metric", font=F_BOLD); C(ws, "C57", "$/SF", font=F_BOLD)
    C(ws, "A58", f"Newest comp ({primary}) avg $/SF")
    C(ws, "C58", f"=IFERROR(AVERAGEIFS('RD Condos'!$E:$E,'RD Condos'!$K:$K,\"{primary}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF2)
    C(ws, "A59", "All-comp avg $/SF"); C(ws, "C59", "=IFERROR(AVERAGEIFS('RD Condos'!$AN:$AN,'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF2)
    C(ws, "A60", f"Oldest comp ({oldest}) avg $/SF")
    C(ws, "C60", f"=IFERROR(AVERAGEIFS('RD Condos'!$E:$E,'RD Condos'!$K:$K,\"{oldest}\",'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF2)
    C(ws, "A61", "Observed vintage premium — newest / all-comp - 1"); C(ws, "C61", "=IFERROR(C58/C59-1,\"\")", nf=NF_PCT)
    C(ws, "A62", "Observed vintage premium — newest / oldest - 1");   C(ws, "C62", "=IFERROR(C58/C60-1,\"\")", nf=NF_PCT)
    C(ws, "A63", "C4 subject premium (sits within observed range)");  C(ws, "C63", "=C4", nf=NF_PCT)

    # ---- MIX BASIS (rows 65-72)
    C(ws, "A65", "MIX BASIS (live) — comp-set bed distribution vs estimated subject mix", font=F_BOLD)
    for col, t in zip("ABCD", ["Bucket","Comp #","Comp share","Subject est."]):
        C(ws, f"{col}66", t, font=F_BOLD)
    C(ws, "A67", "1-Bed (H=1, den-incl)"); C(ws, "B67", "=COUNTIFS('RD Condos'!$H:$H,\"1\",'RD Condos'!$A:$A,1)")
    C(ws, "C67", "=IFERROR(B67/$B$70,\"\")", nf=NF_PCT); C(ws, "D67", "=H2", nf=NF_PCT)
    C(ws, "A68", "2-Bed (H=2, incl 2+den)"); C(ws, "B68", "=COUNTIFS('RD Condos'!$H:$H,\"2\",'RD Condos'!$A:$A,1)")
    C(ws, "C68", "=IFERROR(B68/$B$70,\"\")", nf=NF_PCT); C(ws, "D68", "=H3", nf=NF_PCT)
    C(ws, "A69", "3-Bed (H=3)"); C(ws, "B69", "=COUNTIFS('RD Condos'!$H:$H,\"3\",'RD Condos'!$A:$A,1)")
    C(ws, "C69", "=IFERROR(B69/$B$70,\"\")", nf=NF_PCT); C(ws, "D69", "=H4", nf=NF_PCT)
    C(ws, "A70", "Total included"); C(ws, "B70", "=B67+B68+B69")
    C(ws, "A71", "1+Den share (Q=1+1)"); C(ws, "B71", "=COUNTIFS('RD Condos'!$Q:$Q,\"1+1\",'RD Condos'!$A:$A,1)")
    C(ws, "C71", "=IFERROR(B71/$B$70,\"\")", nf=NF_PCT); C(ws, "D71", "=T4", nf=NF_PCT)

    # ---- APARTMENT PREMIUM BASIS (rows 74-75)
    has_apt = len(d.get("apartment_rows", [])) > 0
    C(ws, "A74", "APARTMENT PREMIUM BASIS (C3)" + ("" if has_apt else " — N/A this set"), font=F_BOLD)
    if has_apt:
        C(ws, "A75", "Derive C3 from condo ÷ apartment $/SF spread over comparable vintages:")
        C(ws, "B75", "=IFERROR(AVERAGEIFS('RD Condos'!$E:$E,'RD Condos'!$A:$A,1)/AVERAGEIFS('RD Apartments'!$E:$E,'RD Apartments'!$A:$A,1)-1,\"\")", nf=NF_PCT)
    else:
        C(ws, "A75", "No purpose-built apartment comps selected, so C3 is unused (0). If apartment "
                     "comps are added, derive C3 from the condo ÷ apartment $/SF spread or cite a "
                     "named market study — never a bare %.", font=F_GREY)

    # ---- comp building rollup (placed at bottom; display-only, not cross-referenced)
    base = 78
    C(ws, f"A{base}", "Comp building rollup (included leases)", font=F_BOLD)
    for col, t in zip("ABCDEFG", ["Building","Address","Built","Incl.","Avg SF","Avg Rent","Avg $/SF"]):
        C(ws, f"{col}{base+1}", t, font=F_HDR, fill=FILL_HDR)
    blds = d["comp_buildings"]
    for i, b in enumerate(blds):
        r = base + 2 + i
        C(ws, f"A{r}", b.get("name"))
        C(ws, f"B{r}", b.get("address"))
        C(ws, f"C{r}", f"Built {b.get('yr_built')}")
        C(ws, f"D{r}", f"=COUNTIFS('RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1)")
        C(ws, f"E{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$B:$B,'RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"F{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$AB:$AB,'RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1),\"\")", nf=NF_INT)
        C(ws, f"G{r}", f"=IFERROR(AVERAGEIFS('RD Condos'!$E:$E,'RD Condos'!$K:$K,$A{r},'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF)
    rall = base + 2 + len(blds)
    C(ws, f"A{rall}", "ALL COMPS", fill=FILL_BAND)
    C(ws, f"D{rall}", "=SUM(D%d:D%d)" % (base+2, base+1+len(blds)), fill=FILL_BAND)
    C(ws, f"E{rall}", "=IFERROR(AVERAGEIFS('RD Condos'!$B:$B,'RD Condos'!$A:$A,1),\"\")", nf=NF_INT, fill=FILL_BAND)
    C(ws, f"F{rall}", "=IFERROR(AVERAGEIFS('RD Condos'!$AB:$AB,'RD Condos'!$A:$A,1),\"\")", nf=NF_INT, fill=FILL_BAND)
    C(ws, f"G{rall}", "=IFERROR(AVERAGEIFS('RD Condos'!$AN:$AN,'RD Condos'!$A:$A,1),\"\")", nf=NF_PSF, fill=FILL_BAND)
    return ws

# ---------------------------------------------------------------- Output
BEDGRP = [("Studio", "L", "0", "H"), ("1-Bedroom", "P", "1", "Q"),
          ("1-Bedroom + Den", "T", "1+1", "Q"), ("2-Bedroom", "X", "2", "H"),
          ("3-Bedroom", "Z" and "X" and "AB", "3", "H")]
# group anchor cols: Studio L, 1BR P, 1+Den T, 2BR X, 3BR AB  (each spans 4 cols)
GRP = [("Studio Units","L","0","H"),("1-Bedroom Units","P","1","Q"),
       ("1-Bedroom + Den Units","T","1+1","Q"),("2-Bedroom Units","X","2","H"),
       ("3-Bedroom Units","AB","3","H")]

def col_off(anchor, n):
    """anchor column letter offset by n (0..3): count/size/rent/psf."""
    from openpyxl.utils import column_index_from_string
    return get_column_letter(column_index_from_string(anchor) + n)

def build_output(wb, d):
    ws = wb.create_sheet("Output")
    ws.sheet_view.showGridLines = False
    widths(ws, {"B":34,"C":13,"D":13,"E":11,"F":13,"G":13,"H":13,"I":14,"K":30,
                "L":11,"M":11,"N":11,"O":11,"P":11,"Q":11,"R":11,"S":11,"T":11,"U":11,
                "V":11,"W":11,"X":11,"Y":11,"Z":11,"AA":11,"AB":11,"AC":11,"AD":11,"AE":11})
    subj = d["subject"]; blds = d["comp_buildings"]; nb = len(blds)
    primary = subj["primary_building"]; rd = "'RD Condos'"

    C(ws, "B2", f"{subj['title_full']} — RENTAL COMP SUMMARY (Condos view)", font=F_TITLE)
    for col, t in zip("BCDEFGHI", ["Building","First Occupancy","Avg.Lease Date","Vacancy",
                                   "Transaction Count","Avg Suite Size","Adj. Avg. Rent","Avg. Net Rent PSF"]):
        C(ws, f"{col}3", t, font=F_HDR, fill=FILL_HDR)
    C(ws, "B5", subj.get("group_label","Comps"), font=F_BOLD, fill=FILL_LBLUE)
    for i, b in enumerate(blds):
        r = 6 + i; bs = 4 + i
        C(ws, f"B{r}", b.get("name"))
        C(ws, f"C{r}", f"='Building Summary'!E{bs}")
        C(ws, f"D{r}", f"=IFERROR(AVERAGEIFS({rd}!$J:$J,{rd}!$K:$K,$B{r},{rd}!$A:$A,1),\"\")", nf=NF_DATE)
        C(ws, f"F{r}", f"='Building Summary'!J{bs}")
        C(ws, f"G{r}", f"='Building Summary'!K{bs}", nf=NF_INT)
        C(ws, f"H{r}", f"=IFERROR(AVERAGEIFS({rd}!$D:$D,{rd}!$K:$K,$B{r},{rd}!$A:$A,1),\"\")", nf=NF_USD)
        C(ws, f"I{r}", f"=IFERROR(AVERAGEIFS({rd}!$F:$F,{rd}!$K:$K,$B{r},{rd}!$A:$A,1),\"\")", nf=NF_USD2)
    avg = 6 + nb; prem = avg+1; impl = avg+2; subjr = avg+4
    exc = avg+6; wavg = exc+4; wprem = wavg+1; wimpl = wavg+2
    if wimpl >= 30:
        raise SystemExit(f"Too many comp buildings ({nb}) for the fixed Output layout "
                         f"(footer would collide with the recommendation fold at row 30). "
                         f"Group buildings or extend the layout.")
    C(ws, f"B{avg}", "Average", font=F_BOLD)
    C(ws, f"F{avg}", f"=COUNTIFS({rd}!$A:$A,1)")
    C(ws, f"G{avg}", f"=IFERROR(AVERAGEIFS({rd}!$B:$B,{rd}!$A:$A,1),\"\")", nf=NF_INT)
    C(ws, f"H{avg}", f"=IFERROR(AVERAGEIFS({rd}!$D:$D,{rd}!$A:$A,1),\"\")", nf=NF_USD)
    C(ws, f"I{avg}", f"=IFERROR(AVERAGEIFS({rd}!$F:$F,{rd}!$A:$A,1),\"\")", nf=NF_USD2)
    C(ws, f"B{prem}", "PBR Premium"); C(ws, f"H{prem}", "=Data_Summary!C4", nf=NF_PCT); C(ws, f"I{prem}", "=Data_Summary!C4", nf=NF_PCT)
    C(ws, f"B{impl}", "Implied Untrended Rent", fill=FILL_LBLUE)
    C(ws, f"G{impl}", f"=G{avg}", nf=NF_INT); C(ws, f"H{impl}", f"=IFERROR(H{avg}*(1+H{prem}),\"\")", nf=NF_USD)
    C(ws, f"I{impl}", f"=IFERROR(I{avg}*(1+I{prem}),\"\")", nf=NF_USD2)
    C(ws, f"B{subjr}", "Subject Site (Untrended Rent)", fill=FILL_CREAM)
    C(ws, f"G{subjr}", f"=G{impl}", nf=NF_INT); C(ws, f"H{subjr}", f"=H{impl}", nf=NF_USD); C(ws, f"I{subjr}", f"=I{impl}", nf=NF_USD2)
    C(ws, f"B{exc}", "Other Excluded", fill=FILL_LBLUE)
    for k, line in enumerate(d.get("other_excluded_prose","").split("\n")):
        C(ws, f"B{exc+1+k}", line, font=F_GREY)
    C(ws, f"B{wavg}", "Weighted Average (all comps)", font=F_BOLD)
    C(ws, f"F{wavg}", f"=COUNTIFS({rd}!$A:$A,1)")
    C(ws, f"G{wavg}", f"=IFERROR(AVERAGEIFS({rd}!$B:$B,{rd}!$A:$A,1),\"\")", nf=NF_INT)
    C(ws, f"H{wavg}", f"=IFERROR(AVERAGEIFS({rd}!$D:$D,{rd}!$A:$A,1),\"\")", nf=NF_USD)
    C(ws, f"I{wavg}", f"=IFERROR(AVERAGEIFS({rd}!$F:$F,{rd}!$A:$A,1),\"\")", nf=NF_USD2)
    C(ws, f"B{wprem}", "PBR Premium"); C(ws, f"H{wprem}", "=Data_Summary!C4", nf=NF_PCT); C(ws, f"I{wprem}", "=Data_Summary!C4", nf=NF_PCT)
    C(ws, f"B{wimpl}", "Implied Untrended Rent", fill=FILL_LBLUE)
    C(ws, f"H{wimpl}", f"=IFERROR(H{wavg}*(1+H{wprem}),\"\")", nf=NF_USD)
    C(ws, f"I{wimpl}", f"=IFERROR(I{wavg}*(1+I{wprem}),\"\")", nf=NF_USD2)

    # ---- right unit-mix grid (fixed rows 27-46) + headers
    for label, anc, _, _ in GRP:
        C(ws, f"{anc}27", label, font=F_BOLD, fill=FILL_LBLUE)
    C(ws, "K28", "Building", font=F_HDR, fill=FILL_HDR)
    for _, anc, _, _ in GRP:
        for n, t in enumerate(["Transaction\nCount","Avg. Suite\nSize","Adj. Avg.\nRent","Avg. Net\nRent PSF"]):
            C(ws, f"{col_off(anc,n)}28", t, font=F_HDR, fill=FILL_HDR, align=Alignment(wrap_text=True))
    def bed_formulas(row, anc, key, kcol, extra=""):
        # count/size/rent/psf for a bed group; kcol is H (beds num) or Q (beds str)
        crit = f"{rd}!${kcol}:${kcol},\"{key}\""
        C(ws, f"{col_off(anc,0)}{row}", f"=COUNTIFS({crit},{rd}!$A:$A,1{extra})")
        C(ws, f"{col_off(anc,1)}{row}", f"=IFERROR(AVERAGEIFS({rd}!$B:$B,{crit},{rd}!$A:$A,1{extra}),\"\")", nf=NF_INT)
        C(ws, f"{col_off(anc,2)}{row}", f"=IFERROR(AVERAGEIFS({rd}!$D:$D,{crit},{rd}!$A:$A,1{extra}),\"\")", nf=NF_USD)
        C(ws, f"{col_off(anc,3)}{row}", f"=IFERROR(AVERAGEIFS({rd}!$F:$F,{crit},{rd}!$A:$A,1{extra}),\"\")", nf=NF_USD2)
    # primary building row (29) and Weighted Avg row (31)
    C(ws, "K29", primary)
    for _, anc, key, kcol in GRP:
        bed_formulas(29, anc, key, kcol, extra=f",{rd}!$K:$K,\"{primary}\"")
    C(ws, "K31", "Weighted Avg.")
    for _, anc, key, kcol in GRP:
        bed_formulas(31, anc, key, kcol)
    # Low / High (array MIN/MAX over included rows of that bed group) — bounded to data extent
    n_condo = len(d["condo_rows"]); last = n_condo + 1
    C(ws, "K32", "Low"); C(ws, "K33", "High")
    for _, anc, key, kcol in GRP:
        rng = f"{rd}!${kcol}$2:${kcol}${last}"
        frng = f"{rd}!$F$2:$F${last}"
        arng = f"{rd}!$A$2:$A${last}"
        C(ws, f"{col_off(anc,3)}32", f"=IFERROR(SUMPRODUCT(MIN(IF((({rng}=\"{key}\")*({arng}=1)),{frng}))),\"\")", nf=NF_USD2)
        C(ws, f"{col_off(anc,3)}33", f"=IFERROR(SUMPRODUCT(MAX(IF((({rng}=\"{key}\")*({arng}=1)),{frng}))),\"\")", nf=NF_USD2)
    C(ws, "K35", "PBR Premium")
    for _, anc, _, _ in GRP:
        C(ws, f"{col_off(anc,3)}35", "=Data_Summary!$C$4", nf=NF_USD2)
    C(ws, "K36", "Weighted Avg. (×(1+PBR))")
    for _, anc, _, _ in GRP:
        rent = col_off(anc,2); psf = col_off(anc,3)
        C(ws, f"{rent}36", f"=IFERROR({rent}31*(1+${psf}$35),\"\")", nf=NF_USD)
        C(ws, f"{psf}36",  f"=IFERROR({psf}31*(1+{psf}35),\"\")", nf=NF_USD2)
    C(ws, "K37", "Weighted Avg. Building Total (Comps)")
    C(ws, "L37", "=IFERROR((R31*Data_Summary!$T$3+V31*Data_Summary!$T$4+Z31*Data_Summary!$T$5+AD31*Data_Summary!$T$6)*Data_Summary!$T$2,\"\")", nf=NF_USD)
    C(ws, "K39", "Subject Property (Using Comp $ Rent)", font=F_BOLD)
    for _, anc, _, _ in GRP:
        rent = col_off(anc,2); C(ws, f"{rent}39", f"=IFERROR({rent}36,\"\")", nf=NF_USD)
    C(ws, "K40", "Subject Property Rent Building Total")
    C(ws, "L40", "=IFERROR((R39*Data_Summary!$T$3+V39*Data_Summary!$T$4+Z39*Data_Summary!$T$5+AD39*Data_Summary!$T$6)*Data_Summary!$T$2,\"\")", nf=NF_USD)
    C(ws, "K42", "Subject Property (Using Comp PSF Rent)")
    for _, anc, _, _ in GRP:
        rent = col_off(anc,2); psf = col_off(anc,3); size = col_off(anc,1)
        C(ws, f"{rent}42", f"=IFERROR({psf}36*{size}31,\"\")", nf=NF_USD)
    C(ws, "K43", "Subject Property PSF Rent Building Total")
    C(ws, "L43", "=IFERROR((R42*Data_Summary!$T$3+V42*Data_Summary!$T$4+Z42*Data_Summary!$T$5+AD42*Data_Summary!$T$6)*Data_Summary!$T$2,\"\")", nf=NF_USD)
    C(ws, "K45", "Subject Property Rent", fill=FILL_LBLUE)
    for _, anc, _, _ in GRP:
        rent = col_off(anc,2); C(ws, f"{rent}45", f"=IFERROR(AVERAGE({rent}39,{rent}42),\"\")", nf=NF_USD)
    C(ws, "K46", "Subject Property Rent Building Total")
    C(ws, "L46", "=IFERROR((R45*Data_Summary!$T$3+V45*Data_Summary!$T$4+Z45*Data_Summary!$T$5+AD45*Data_Summary!$T$6)*Data_Summary!$T$2,\"\")", nf=NF_USD)

    # ---- regression / coefficients
    C(ws, "K48", "Regression:", fill=FILL_REG)
    C(ws, "K49", "Coefficients:")
    C(ws, "K50", "Total Building coefficient", fill=FILL_REG); C(ws, "L50", 0)
    C(ws, "M50", "← optional external regression input (Predictor); 0 = unused", nf=NF_INT)
    C(ws, "K52", "Subject Property Rent (Increase vs comps)"); C(ws, "L52", "=IFERROR(L46/L37-1,\"\")", nf=NF_PCT)
    C(ws, "K53", "Subject Property Rent (Building Total)", fill=FILL_REG); C(ws, "L53", "=L46")

    # ---- TRREB table
    tr = d.get("trreb", {})
    C(ws, "R56", "Unadjusted"); C(ws, "Z56", "Unadjusted"); C(ws, "AD56", "Unadjusted")
    C(ws, "K57", f"TRREB Data ({tr.get('quarter_label','')})")
    C(ws, "R57", "Rent (1BR)"); C(ws, "Z57", "Rent (2BR)"); C(ws, "AD57", "Rent (3BR)")
    C(ws, "K58", tr.get("district_label","")); C(ws, "P58", "← TRREB (district / City / GTA YoY), read this session", font=F_GREY)
    if tr:
        C(ws, "R58", tr["district"]["1BR"], font=F_BLUE); C(ws, "Z58", tr["district"]["2BR"], font=F_BLUE); C(ws, "AD58", tr["district"]["3BR"], font=F_BLUE)
        C(ws, "K59", "City of Toronto")
        C(ws, "R59", tr["city"]["1BR"], font=F_BLUE); C(ws, "Z59", tr["city"]["2BR"], font=F_BLUE); C(ws, "AD59", tr["city"]["3BR"], font=F_BLUE)
        C(ws, "K60", "YoY Change (GTA)")
        C(ws, "R60", tr["yoy"]["1BR"], nf=NF_PCT, font=F_BLUE); C(ws, "Z60", tr["yoy"]["2BR"], nf=NF_PCT, font=F_BLUE); C(ws, "AD60", tr["yoy"]["3BR"], nf=NF_PCT, font=F_BLUE)

    # ---- recommendation fold (rows 30-48, cols B-D)
    C(ws, "B30", "★ RECOMMENDATION — RECOMMENDED SUBJECT RENT", font=F_BOLD)
    C(ws, "B31", "Recommended subject $/SF"); C(ws, "D31", "=Data_Summary!$C$34", nf=NF_PSF2)
    C(ws, "B32", "Conservative (all-comp + premium) $/SF"); C(ws, "D32", "=Data_Summary!$C$35", nf=NF_PSF2)
    C(ws, "B33", f"Raw comp $/SF ({primary}, no premium)"); C(ws, "D33", "=Data_Summary!$C$31", nf=NF_PSF2)
    C(ws, "B34", "Plus — new-build / lease-up premium"); C(ws, "D34", "=Data_Summary!$C$4", nf=NF_PCT)
    C(ws, "B36", "Recommended monthly rent by suite type", font=F_BOLD)
    C(ws, "B37", "Suite", font=F_BOLD); C(ws, "C37", "Avg SF", font=F_BOLD); C(ws, "D37", "Rec. Rent", font=F_BOLD)
    C(ws, "B38", "1-Bed / 1+Den"); C(ws, "C38", "=Data_Summary!C26", nf="0"); C(ws, "D38", "=Data_Summary!F26", nf=NF_USD)
    C(ws, "B39", "2-Bed"); C(ws, "C39", "=Data_Summary!C27", nf="0"); C(ws, "D39", "=Data_Summary!F27", nf=NF_USD)
    C(ws, "B40", "3-Bed"); C(ws, "C40", "=Data_Summary!C28", nf="0"); C(ws, "D40", "=Data_Summary!F28", nf=NF_USD)
    C(ws, "B42", "Custom suite — enter size (sq ft) →"); C(ws, "D42", 700, fill=FILL_YEL)
    C(ws, "B43", "Implied rent @ recommended subject $/SF", font=F_BOLD); C(ws, "D43", "=IFERROR(D42*Data_Summary!$C$34,\"\")", nf=NF_USD)
    C(ws, "B44", "Implied rent @ raw comp $/SF (no premium)", font=F_BOLD); C(ws, "D44", "=IFERROR(D42*Data_Summary!$C$31,\"\")", nf=NF_USD)
    C(ws, "B46", "NOTES & CONTEXT", font=F_BOLD)
    notes = subj.get("notes", [])
    notes = notes + ["Confidence fills (Raw Data only): green = leased in-window + SF verified per-unit · "
                     "cream = plate-verified / older / minor caveat / apartment asking-only · orange = excluded "
                     "partial / bracket-only. No colour fills on summary/Output cells."]
    for k, line in enumerate(notes):
        C(ws, f"B{47+k}", line, font=F_GREY)
    return ws

# ---------------------------------------------------------------- recalc
def recalc_in_place(path):
    """Bake computed formula VALUES into the .xlsx so it displays correctly in ANY viewer.

    openpyxl writes formula strings but no cached results, so a non-Excel viewer
    (macOS Quick Look / Numbers, Google Sheets, file-preview panes) shows every
    computed cell blank until something recalculates. Excel does recalc on open
    (fullCalcOnLoad), but other viewers do not. This runs the finished workbook through
    LibreOffice headless with "always recalculate on load" so the results are cached in
    the file. Graceful no-op if LibreOffice isn't installed (the formulas remain; Excel
    will still recalc on open). NOTE: this trades the byte-for-byte determinism of the
    pure-openpyxl output for display-ready cached values.
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        print("  recalc: LibreOffice not found — left formulas uncached (Excel recalcs on open)")
        return False
    prof = tempfile.mkdtemp(prefix="lo_prof_")
    os.makedirs(os.path.join(prof, "user"), exist_ok=True)
    with open(os.path.join(prof, "user", "registrymodifications.xcu"), "w") as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n'
                '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
                'xmlns:xs="http://www.w3.org/2001/XMLSchema" '
                'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
                ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
                '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
                ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
                '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
                '</oor:items>\n')
    outdir = tempfile.mkdtemp(prefix="lo_out_")
    try:
        subprocess.run([soffice, "-env:UserInstallation=file://%s" % prof,
                        "--headless", "--calc",
                        "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                        "--outdir", outdir, path],
                       check=True, capture_output=True, timeout=180)
    except Exception as e:
        print("  recalc: LibreOffice run failed (%s) — left formulas uncached" % e)
        return False
    produced = os.path.join(outdir, os.path.basename(path))
    if os.path.exists(produced):
        shutil.copyfile(produced, path)
        print("  recalc: formula values cached via LibreOffice")
        return True
    print("  recalc: LibreOffice produced no output — left formulas uncached")
    return False

# ---------------------------------------------------------------- main
def build(data, out_path, recalc=True):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_output(wb, data)
    build_building_summary(wb, data["subject"], data["comp_buildings"])
    build_data_summary(wb, data)
    build_rd(wb, "RD Condos", data["condo_rows"], "Condo")
    build_rd(wb, "RD Apartments", data["apartment_rows"], "Apartment")
    fp_note = ("NOTE — one row per distinct VIPcondos plan opened this session. Subject plans "
               "tagged '(SUBJECT)'.") if data["floor_plans"] else \
              ("No VIPcondos plans opened this session — all comp SF from condos.ca registered areas (Route B).")
    build_floor_plans(wb, data["floor_plans"], fp_note)
    # workbook order: Output, Building Summary, Data_Summary, RD Condos, RD Apartments, Floor Plans
    order = ["Output","Building Summary","Data_Summary","RD Condos","RD Apartments","Floor Plans"]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    # size every column to its content (replaces the hand-tuned widths; no wrap)
    for ws in wb.worksheets:
        autofit(ws)
    # pin document timestamps so identical input -> byte-identical output (deterministic)
    fixed = datetime.datetime(2020, 1, 1)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.save(out_path)
    # bake cached formula values so the file displays in any viewer (not just Excel)
    if recalc:
        recalc_in_place(out_path)
    return out_path

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: python build_workbook.py comp_data.json [out.xlsx]")
    data = json.load(open(sys.argv[1], encoding="utf-8"))
    out = sys.argv[2] if len(sys.argv) > 2 else f"{data['subject']['name']} Rental Comps _vACTIVE.xlsx"
    path = build(data, out)
    print("WROTE", path)
