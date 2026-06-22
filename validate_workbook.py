#!/usr/bin/env python3
"""
validate_workbook.py — conformance + recalc + numpy tie-out for a generated v2 workbook.

    python validate_workbook.py "<workbook>.xlsx" comp_data.json

Checks, in order:
  1. STRUCTURE  — 6 sheets in order; no Subject & Conclusion / INPUTS; column widths set on
                  every sheet; gridlines hidden; freeze panes on RD/Floor Plans; confidence
                  fills only on RD cols B/E.
  2. RECALC     — evaluate every formula (via the `formulas` engine, on a copy whose
                  whole-column refs are bounded to the data extent) and assert ZERO error cells.
  3. TIE-OUT    — independently recompute the key numbers in numpy from comp_data and confirm
                  the workbook's evaluated cells match.
  4. COVERAGE   — every leased-in-window unit (leased rent present AND date >= C1) either
                  carries a verified SF or a documented SF-resolution exhaustion record
                  (`sf_blank`: condosca_page + keyplate + outcome). A silent blank — a unit
                  with the rent but no SF and no record of which routes were tried — FAILS.
                  Prints the SF-resolution coverage ratio (previously invisible).

Exit code 0 = pass. Non-zero = at least one check failed (details printed).
"""
import sys, json, re, tempfile, os, warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

ORDER = ["Output", "Building Summary", "Data_Summary", "RD Condos", "RD Apartments", "Floor Plans"]
CONF = {"FFE2EFDA", "FFFFF2CC", "FFFCE4D6"}
fails = []
def check(cond, msg):
    if not cond: fails.append(msg)
    print(("  ok  " if cond else "  FAIL ") + msg)

def structure(path):
    print("STRUCTURE")
    wb = openpyxl.load_workbook(path)
    check(wb.sheetnames == ORDER, f"sheets == {ORDER} (got {wb.sheetnames})")
    check("Subject & Conclusion" not in wb.sheetnames, "no 'Subject & Conclusion' sheet")
    for s in wb.sheetnames:
        ws = wb[s]
        nW = sum(1 for v in ws.column_dimensions.values() if v.width)
        check(nW >= 4, f"[{s}] column widths set (got {nW})")
        check(ws.sheet_view.showGridLines is False, f"[{s}] gridlines hidden")
    for s in ("RD Condos", "RD Apartments", "Floor Plans"):
        check(wb[s].freeze_panes == "A2", f"[{s}] freeze panes A2")
    # confidence fills: on RD sheets only B/E may carry any of the 3 confidence colours;
    # the pure data-confidence colours (green E2EFDA / orange FCE4D6) must not appear on any
    # summary sheet. (Cream FFF2CC doubles as a legitimate Output label highlight, e.g. the
    # "Subject Site" row, exactly as in the reference workbook — so it is allowed there.)
    DATA_CONF = {"FFE2EFDA", "FFFCE4D6"}
    stray = []
    for s in wb.sheetnames:
        ws = wb[s]
        for row in ws.iter_rows():
            for c in row:
                fg = getattr(c.fill.fgColor, "rgb", None) if c.fill else None
                if s in ("RD Condos", "RD Apartments"):
                    if fg in CONF and c.column_letter not in ("B", "E"):
                        stray.append(f"{s}!{c.coordinate}")
                elif fg in DATA_CONF:
                    stray.append(f"{s}!{c.coordinate}")
    check(not stray, f"data-confidence fills confined to RD B/E (stray: {stray[:5]})")
    # INPUTS table removed
    ds = wb["Data_Summary"]
    has_inputs = any(str(ds.cell(r, 1).value or "").strip().upper().startswith("INPUTS")
                     for r in range(1, ds.max_row + 1))
    check(not has_inputs, "no INPUTS table on Data_Summary")
    return wb

def bound_ranges(path, counts):
    """Rewrite whole-column refs ('Sheet'!$A:$A) to bounded ranges so `formulas` is fast."""
    wb = openpyxl.load_workbook(path)
    pat = re.compile(r"('?[A-Za-z0-9 ]+'?)!\$([A-Z]{1,3}):\$([A-Z]{1,3})")
    def last_for(sheetref):
        nm = sheetref.strip("'")
        return counts.get(nm, 2000)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    def repl(m):
                        sh, c1, c2 = m.group(1), m.group(2), m.group(3)
                        last = last_for(sh)
                        return f"{sh}!${c1}$2:${c2}${last}"
                    c.value = pat.sub(repl, c.value)
    tmp = os.path.join(tempfile.gettempdir(), "_bounded_eval.xlsx")
    wb.save(tmp)
    return tmp

def recalc(path, counts):
    print("RECALC (bounded-range eval)")
    import formulas
    tmp = bound_ranges(path, counts)
    xl = formulas.ExcelModel().loads(tmp).finish()
    sol = xl.calculate()
    errs = {}
    vals = {}
    for k, v in sol.items():
        try: val = v.value[0, 0]
        except Exception:
            try: val = v.value
            except Exception: val = None
        addr = k.upper().split("!")[-1].replace("'", "")
        vals[k.upper()] = val
        ERR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!")
        if isinstance(val, str) and any(val.startswith(t) for t in ERR_TOKENS):
            errs[k] = val
    check(len(errs) == 0, f"zero formula errors (got {len(errs)}: {list(errs.items())[:6]})")
    return sol

def cell(sol, sheet, addr):
    key = f"'[{os.path.basename('_bounded_eval.xlsx')}]{sheet}'!{addr}".upper()
    for k, v in sol.items():
        ku = k.upper()
        if ku.endswith(f"]{sheet}'!{addr}".upper()) or ku.endswith(f"{sheet}'!{addr}".upper()):
            try: return v.value[0, 0]
            except Exception: return v.value
    return None

def tieout(sol, data):
    print("TIE-OUT (numpy recomputation)")
    import numpy as np
    lev = data["levers"]; C2 = lev["parking_adj"]["value"]; C3 = lev["apt_premium"]["value"]
    df = lev["date_filter"]
    primary = data["subject"]["primary_building"]
    rows = data["condo_rows"]
    def inc(r):
        return bool(r.get("sqft")) and bool(r.get("leased_rent")) and str(r.get("date","")) >= df
    E = lambda r: r["leased_rent"] / r["sqft"]
    Hd = lambda r: str(r.get("beds",""))[:1]
    incl = [r for r in rows if inc(r)]
    def avgE(sel):
        xs = [E(r) for r in sel]
        return float(np.mean(xs)) if xs else None
    c58 = avgE([r for r in incl if r["building"] == primary])
    c59 = avgE(incl)  # AN == E for condos
    # by-bed primary E (rows 26/27/28)
    e26 = avgE([r for r in incl if r["building"] == primary and Hd(r) == "1"])
    e27 = avgE([r for r in incl if r["building"] == primary and Hd(r) == "2"])
    e28 = avgE([r for r in incl if r["building"] == primary and Hd(r) == "3"])
    smm = lev["subject_mix_manual"]; H2, H3, H4 = smm["1BR"], smm["2BR"], smm["3BR"]
    c31 = (e26 or 0)*H2 + (e27 or 0)*H3 + (e28 or 0)*H4
    oldest = data["subject"].get("oldest_building", primary)
    c60 = avgE([r for r in incl if r["building"] == oldest])
    c61 = c58/c59 - 1 if c59 else 0
    c62 = c58/c60 - 1 if c60 else 0
    sp = lev.get("subject_premium")
    if abs(c61) < 1e-4 and abs(c62) < 1e-4:   # single-/equal-vintage comp set: no spread to derive
        c4 = (sp["value"] if sp else 0.0)
    else:
        c4 = (c61 + c62)/2
    c34 = c31*(1+c4)
    exp = {"DATA_SUMMARY!C58": c58, "DATA_SUMMARY!C59": c59, "DATA_SUMMARY!C60": c60,
           "DATA_SUMMARY!C31": c31, "DATA_SUMMARY!C4": c4, "DATA_SUMMARY!C34": c34}
    for addr, want in exp.items():
        sheet, a = addr.split("!")
        got = cell(sol, "DATA_SUMMARY", a)
        try: gotf = float(got)
        except Exception: gotf = None
        okc = (gotf is not None and want is not None and abs(gotf - want) < max(0.01, abs(want)*0.001))
        check(okc, f"{addr}: workbook={got!r} vs numpy={None if want is None else round(want,4)}")
    print(f"    [recommended subject $/SF = {round(c34,3)}; all-comp $/SF = {round(c59,3)}; "
          f"included rows = {len(incl)}]")

def coverage(data):
    """A leased-in-window unit must have an SF or a documented exhaustion record — a silent
    blank (bounded effort masquerading as 'cannot verify') fails. This is the negation of the
    Include predicate on the SF term only: leased rent present, date >= C1, but no SF."""
    print("SF-COVERAGE (leased-in-window units: verified SF, or a documented sf_blank)")
    df = data["levers"]["date_filter"]
    REQ = ("condosca_page", "keyplate", "outcome")
    rows = ([("RD Condos", r) for r in data.get("condo_rows", [])]
            + [("RD Apartments", r) for r in data.get("apartment_rows", [])])
    leased_inwin = [(s, r) for s, r in rows
                    if r.get("leased_rent") and str(r.get("date", "")) >= df]
    blanks = [(s, r) for s, r in leased_inwin if not r.get("sqft")]
    undocumented = []
    for s, r in blanks:
        sb = r.get("sf_blank")
        ok = isinstance(sb, dict) and all(str(sb.get(k, "") or "").strip() for k in REQ)
        if not ok:
            undocumented.append(f"{s}!unit {r.get('unit', '?')}")
    msg = ("every leased-in-window blank-SF row carries a complete sf_blank "
           "(condosca_page + keyplate + outcome)")
    if undocumented:
        more = f" (+{len(undocumented) - 8} more)" if len(undocumented) > 8 else ""
        msg += f" — MISSING on: {undocumented[:8]}{more}"
    check(not undocumented, msg)
    resolved = len(leased_inwin) - len(blanks)
    documented = len(blanks) - len(undocumented)
    ratio = (100.0 * resolved / len(leased_inwin)) if leased_inwin else 100.0
    print(f"    [leased in-window = {len(leased_inwin)}; SF resolved = {resolved} "
          f"({round(ratio, 1)}%); blank w/ exhaustion record = {documented}; "
          f"blank undocumented = {len(undocumented)}]")

def main():
    path = sys.argv[1]; data = json.load(open(sys.argv[2], encoding="utf-8"))
    counts = {"RD Condos": len(data["condo_rows"]) + 1,
              "RD Apartments": max(2, len(data["apartment_rows"]) + 1),
              "Floor Plans": len(data["floor_plans"]) + 1}
    structure(path)
    sol = recalc(path, counts)
    tieout(sol, data)
    coverage(data)
    print("\n" + ("PASS — all checks green" if not fails else f"FAIL — {len(fails)} issue(s)"))
    sys.exit(1 if fails else 0)

if __name__ == "__main__":
    main()
